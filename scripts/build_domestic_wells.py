"""Ingest the LWA domestic-well inventory (domestic_wells.xlsx) and
spatial-join each well to our 2027 three-zone polygons for the §5.3
MT-sensitivity feature.

Source: domestic_wells.xlsx  (1,472 wells, WGS84 lat/long columns)

This REPLACES the older cosmo-sourced domestic-well pipeline
(scripts/fetch_cosmo_domestic_wells.py). We derive `our_polygon` by
point-in-polygon against our own 26 three-zone polygons, so the
well->polygon mapping stays consistent with the rest of the dashboard.

We map the workbook columns to the schema the dashboard consumes
(js/main.js reads `DOMESTIC_WELLS` and needs lat/lon, well_bottom_amsl,
local_gse, include):

    wid, lat, lon, install_date, well_depth, local_gse, well_bottom_amsl,
    perf_top_amsl, perf_bot_amsl, include, accuracy

We add (same as the old pipeline):
    our_polygon    — zone_label of the containing 2027 three-zone polygon
                     (point-in-polygon), or None if outside all of them.
    our_mgmt_area  — short name of that polygon (North / Chico / South)

Outputs:
    data/domestic_wells.json         — pruned + spatial-joined
    js/domestic-wells-data.js        — dashboard bundle: `const DOMESTIC_WELLS = [...]`
"""
from __future__ import annotations

import datetime as _dt
import json
import math
from pathlib import Path

import openpyxl
from shapely.geometry import Point, shape

ROOT = Path(__file__).resolve().parent.parent
XLSX_IN = ROOT / "domestic_wells.xlsx"
POLYGONS_GEOJSON = ROOT / "data" / "vina_2027_thiessen_three_zone.geojson"
JSON_OUT = ROOT / "data" / "domestic_wells.json"
JS_OUT = ROOT / "js" / "domestic-wells-data.js"

# Map workbook column header -> output schema field.  lat/long drive geometry;
# `include?` handled separately (header has a trailing '?').
COL_MAP = {
    "WID": "wid",
    "well_depth": "well_depth",
    "GSE": "local_gse",
    "well_bottom": "well_bottom_amsl",   # well-bottom elevation (ft amsl) — the dry-test field
    "perf_top_amsl": "perf_top_amsl",
    "perf_bot_amsl": "perf_bot_amsl",
    "accuracy": "accuracy",
}


def load_polygons():
    """Return [(zone_label, mgmt_area_short, geom)] for our 26 polygons.

    buffer(0) cleans any minor topology issues (slivers / near-zero dangles)
    from the clip without meaningfully changing geometry.
    """
    fc = json.loads(POLYGONS_GEOJSON.read_text())
    out = []
    for feat in fc["features"]:
        props = feat["properties"]
        geom = shape(feat["geometry"])
        if not geom.is_valid:
            geom = geom.buffer(0)
        out.append((props["zone_label"], props.get("mgmt_area"), geom))
    return out


def clean(v):
    """JSON-safe scalar: NaN -> None, datetime -> year int, else pass through."""
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.year
    return v


def open_workbook(path: Path):
    """Load the workbook, tolerating an Excel lock by reading a temp copy.

    Python's open() can't read a file Excel holds open, but an OS-level copy
    (cp on POSIX / GitBash, copy on cmd) opens with FILE_SHARE_READ and works.
    """
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        import subprocess
        import tempfile
        tmp = Path(tempfile.gettempdir()) / f"_{path.name}"
        for cmd in (["cp", str(path), str(tmp)],
                    ["cmd", "/c", "copy", "/y", str(path), str(tmp)]):
            try:
                subprocess.run(cmd, check=True, capture_output=True)
                print(f"  (source locked by Excel — reading temp copy {tmp})")
                return openpyxl.load_workbook(tmp, read_only=True, data_only=True)
            except (FileNotFoundError, subprocess.CalledProcessError):
                continue
        raise SystemExit(
            f"Could not read {path.name} — it appears open in Excel. "
            "Close it and re-run."
        )


def main() -> None:
    wb = open_workbook(XLSX_IN)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header)}
    print(f"Read header from {XLSX_IN.name}: {len(header)} columns")

    polys = load_polygons()
    print(f"Loaded {len(polys)} three-zone polygons for spatial join")

    def get(row, col):
        return row[idx[col]] if col in idx else None

    out = []
    by_polygon: dict[str | None, int] = {}
    n_dropped_geom = 0
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        lat, lon = get(row, "lat"), get(row, "long")
        if lat is None or lon is None:
            n_dropped_geom += 1
            continue
        pt = Point(float(lon), float(lat))

        match_label = match_ma = None
        for zone_label, mgmt_area_short, poly in polys:
            if poly.covers(pt):
                match_label, match_ma = zone_label, mgmt_area_short
                break

        rec = {out_key: clean(get(row, src)) for src, out_key in COL_MAP.items()}
        rec["install_date"] = clean(get(row, "inst_date"))
        rec["lat"] = float(lat)
        rec["lon"] = float(lon)
        inc = get(row, "include?")
        rec["include"] = int(inc) if inc is not None else 1
        rec["our_polygon"] = match_label
        rec["our_mgmt_area"] = match_ma
        out.append(rec)
        by_polygon[match_label] = by_polygon.get(match_label, 0) + 1

    kept_active = sum(1 for r in out if r["include"] == 1)
    print(f"\n  kept {len(out):,} wells (dropped {n_dropped_geom} with no lat/long)")
    print(f"  include=1 (active):      {kept_active}")
    print(f"  outside all 26 polygons: {by_polygon.get(None, 0)}")

    print("\n  by polygon (top 10):")
    items = sorted(((k, v) for k, v in by_polygon.items() if k), key=lambda x: -x[1])
    for k, v in items[:10]:
        print(f"    {k:<24} {v:>5}")

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {JSON_OUT}")

    JS_OUT.parent.mkdir(parents=True, exist_ok=True)
    JS_OUT.write_text(
        "// Auto-generated by scripts/build_domestic_wells.py - do not edit by hand.\n"
        f"// Source: {XLSX_IN.name}\n"
        "// Pruned to dashboard-relevant fields + spatial-joined against our\n"
        "// 2027 three-zone polygons (each well's `our_polygon` field is the\n"
        "// zone_label of the containing 2027 polygon, or null if outside the\n"
        "// 26-polygon coverage). `our_mgmt_area` is the short name (North /\n"
        "// Chico / South) for color/legend purposes.\n\n"
        "const DOMESTIC_WELLS = " + json.dumps(out) + ";\n"
    )
    print(f"Wrote {JS_OUT} ({JS_OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
